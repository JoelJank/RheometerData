import csv
import numpy as np
import matplotlib.pyplot as plt
from scipy.signal import savgol_filter
import math
import os 
import sys
import pandas as pd
from openpyxl import load_workbook
import json

settings_path = "settings/settings.json"

def get_sheetnames(Filepath):
    wb = load_workbook(Filepath, read_only=True, keep_links=False)
    return wb.sheetnames

def get_settings_file(Settings_path):
    with open(Settings_path,'r') as f:
        return json.load(f)

def get_excel_file(Excel_path, Sheet_name):
    file = pd.read_excel(Excel_path, sheet_name = Sheet_name,header = 0,names= ["No","Time","Gap","Normal Force","Normal Stress","Torque","Shear Stress","Rotational Speed"] , skiprows=1)
    mask = file.apply(lambda row: any(isinstance(x, str) for x in row), axis=1)
    file = file[~mask].dropna().reset_index(drop=True)
    return file

def lin_fit(m, n, x):
    return m * x + n


settings = get_settings_file(settings_path)
plt.style.use(settings["plotstyle"])

sheetnames = get_sheetnames(settings["input_file"])
del sheetnames[:settings["skip_sheets"]]
sheetnames = sheetnames[:-settings["ignore_sheets_end"] or None]

processed_data = []
timessavgol_max = []
print("Program started")

for sheet in sheetnames:
    current_data = []
    times_this_sheet = []
    df = get_excel_file(settings["input_file"], sheet)
    time_col = df["Time"].values 
    jumps_where = np.where(np.abs(np.diff(time_col)) > 10)[0] #Determine switch between different consolidation forces
    jumps_end = jumps_where[-1]+jumps_where[0]+1
    jumps = np.append(jumps_where, jumps_end)
    jump_starts = np.array([0])
    for i in range(len(jumps)-1):
        jump_starts = np.append(jump_starts, jumps[i]+1)


    fig, ax = plt.subplots(figsize=(12,5))
    ax2 = ax.twinx()
    ax3 = ax.twinx()
    ax3.spines.right.set_position(("axes",1.1))
    
    
    
    #Normal Stress with mean
    p1, = ax.plot(df["Time"],df["Normal Stress"],marker="o",linestyle="None",markersize=1, color="blue",label="Normal Stress")
    
    for i in range(len(jump_starts)):
        mean = np.mean(df["Normal Stress"][jump_starts[i]:jumps[i]])
        ax.hlines(mean, xmin = df["Time"][jump_starts[i]], xmax = df["Time"][jumps[i]], color="red", linestyle="--")
        current_data.append(mean)
    ax.set_ylabel("Normal Stress [Pa]"); ax.set_xlabel("Time [s]")
    ax.yaxis.label.set_color(p1.get_color())
    ax.tick_params(axis = 'y', colors = p1.get_color())
    ax.set_title(f"{sheet}")
    lines1, labels1 = ax.get_legend_handles_labels()
    
    
    
    #Shear Stress with savgol filter
    p2, = ax2.plot(df["Time"],df["Shear Stress"],marker="o",linestyle = "None", markersize=1, color="orange", label="Shear Stress")
    for i in range(len(jump_starts)):
        sav = savgol_filter(df["Shear Stress"][jump_starts[i]:jumps[i]], settings["savgol_window_length"], settings["savgol_polyorder"])
        p2_sav, = ax2.plot(df["Time"][jump_starts[i]:jumps[i]], sav, color="purple", linestyle = "--")
        sav_max = [np.max(sav), list(sav).index(max(sav))]
        p2_max, = ax2.plot(df["Time"][jump_starts[i]+sav_max[1]], sav[sav_max[1]], marker="x", color = "red", linestyle = "None")
        times_this_sheet.append(df["Time"][jump_starts[i]+sav_max[1]])
        current_data.append(sav[sav_max[1]])
    ax2.set_ylabel("Shear Stress [Pa]")
    ax2.yaxis.label.set_color(p2.get_color())
    ax2.tick_params(axis = 'y', colors = p2.get_color())
    ax2.grid()
    lines2, labels2 = ax2.get_legend_handles_labels()
    
    
    #Gap
    p3, = ax3.plot(df["Time"], df["Gap"], marker = "o", linestyle = "None", markersize = 1, color = "green", label = "Gap[mm]")
    ax3.set_ylabel("Gap [mm]")
    ax3.yaxis.label.set_color(p3.get_color())
    ax3.tick_params(axis = 'y', colors = p3.get_color())
    lines3, labels3 = ax3.get_legend_handles_labels()
    
    #Legend
    ax2.legend(lines1 + lines2 + lines3, labels1+labels2+labels3, loc='upper center', bbox_to_anchor=(0.5, 1.2), ncol=3, markerscale = 6)
    
    plt.savefig(f"{settings['output_folder']}/{sheet}_original_data.png", dpi=300, bbox_inches='tight')
    
    #Save data for mohr-coulomb graph
    current_data = np.array(current_data)
    current_data = np.reshape(current_data, (2, len(jump_starts)))
    processed_data.append(current_data)
    timessavgol_max.append(times_this_sheet)
    
    
    
#Mohr-Coulomb
fit_data= []
plt.figure(figsize=(12,5))
    
for i in range(len(processed_data)):
    plot = plt.plot(processed_data[i][0],processed_data[i][1], linestyle = "None", marker = "x", label = f"{sheetnames[i]}")
    m, n = np.polyfit(processed_data[i][0], processed_data[i][1], 1)
    y_intercept = round(n, 4)
    angle = round(math.degrees(math.atan(m)),4)
    fit_data.append([m, n, y_intercept, angle])
    x = np.linspace(0, max(processed_data[i][0]), 100)
    plt.plot(x, lin_fit(m, n, x), linestyle = "--", color = plot[0].get_color())
    plt.text(0.2,0.95 - i*0.05,
             f"{sheetnames[i]}: y-intercept = {y_intercept}; Angle = {angle}°",
             transform = plt.gca().transAxes,
             fontsize=12,
             verticalalignment='top')
plt.xlabel("Normal Shear [Pa]"); plt.ylabel("Shear Stress [Pa]")
plt.title(f"{settings['title_of_experiment']} Mohr-Coulomb")
plt.legend(loc='upper left')
plt.grid()



plt.savefig(f"{settings['output_folder']}/{settings['title_of_experiment']}_Mohr-Coulomb.png", dpi=300, bbox_inches='tight')

csv_path = f"{settings['output_folder']}/mohr_coulomb.csv"
with open(csv_path, "w", newline='') as csvfile:
    writer = csv.writer(csvfile, delimiter=';')
    writer.writerow([
        "Sheetname", "m", "n", "y-intercept", "Angle",
        "Times_max", "Normal Stress", "Shear Stress max"
    ])
    for i in range(len(fit_data)):
        writer.writerow([
            sheetnames[i],
            round(fit_data[i][0], 3),
            round(fit_data[i][1], 3),
            fit_data[i][2],
            fit_data[i][3],
            ";".join(map(str, timessavgol_max[i])),  # oder spezifisch pro Sheet
            ";".join(map(str, processed_data[i][0])),
            ";".join(map(str, processed_data[i][1]))
        ])

print("Program finished")